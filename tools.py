# functions for converting the bank statements

import os.path
import sys
import json
import pandas as pd
import datetime
import xlsxwriter
import logging
import xlrd
import openpyxl as oxl

budget_titles = []

"""
from __future__ import print_function
import pickle
import os.path
from Google import Create_Service
"""


######################################
# MOVE THESE TO THE CONFIG.JSON FILE
# transaction_expense_start_cell = (5, 2)
# transaction_income_start_cell = (5, 7)
#####################################

def get_google_account_keys_file_from_json(keys_json_file='config.json'):
    # get google account keys
    try:
        with open(keys_json_file) as f:
            return json.load(f)
    except FileNotFoundError as fe:
        logging.error("---NO CONFIG FILE FOUND -- STOPPING EXECUTION---")
        logging.error(fe)
        sys.exit(0)


"""
def write_to_expense_report(filepath,worksheet_name,write_start_row,write_start_col,user_df):
    
    wb = oxl.load_workbook(filepath)

    print("current active worksheet")
    print(wb.sheetnames)

    
    for row_num, row_data in enumerate(user_df):
        for col_num, col_data in enumerate(row_data):
            ws.write(row_num,col_num,col_data)
"""


def write_to_expense_report(workbook_path, summary_df, transactions_df):

    # define workbook object, if the path already exists, form the object using that path
    """
    if os.path.exists(workbook_path):
        
        wb = oxl.load_workbook(workbook_path)
    """

    wb = oxl.Workbook()

    wb.title = workbook_path[workbook_path.rfind("/")+1:len(workbook_path)-1] if workbook_path.rfind("/") != -1 else workbook_path

    # create export report sheets
    wb.create_sheet("Summary")
    wb.create_sheet("Transactions")

    # verify that they were created
    logging.debug("worksheets in your current working xlsx file")
    
    # pip install gspread oauth2clientsheets = wb.sheetnames
    
    logging.debug(wb.sheetnames)


# verification 

def main():
    allKeys = get_google_account_keys_file_from_json()['runtime']
    statement_filepath = allKeys['statement_filepath']

    try:
        statementDF = pd.read_csv("resources/test-export.csv")
    except FileNotFoundError as fe:
        logging.error("File not found! Exiting program")
        logging.error(fe)
        sys.exit(0)
    statementDF["Date"] = pd.to_datetime(statementDF['Date'])

    # filter out the days from the statement
    datedStatementDF = statementDF.loc[
        (statementDF['Date'] > allKeys['start_date']) & (statementDF['Date'] <= allKeys['end_date'])]

    # filter out income from paychecks (need to be more specific for all income transactions, limit to paycheck for now)
    paycheckDF = datedStatementDF.loc[
        (datedStatementDF['Category'] == 'Paycheck') | (datedStatementDF['Category'] == 'Income')]

    # filter out all transactions that aren't paycheck
    expensesDF = \
    datedStatementDF.loc[(datedStatementDF['Category'] != 'Paycheck') & (datedStatementDF['Category'] != 'Income')][
        allKeys['expense_col_names']]

    # get list of unique categories
    categoryList = expensesDF['Category'].unique()

    # write expense df to transactions worksheet
    # write_to_expense_report(allKeys["budget_filepath"],'Transactions',transaction_expense_start_cell[0],transaction_expense_start_cell[1],expensesDF)
    write_to_expense_report(allKeys['expense_report'])

    """
    # test functions to check on the dataframes (need better testing practices)
    print("paycheckDF")
    #print(paycheckDF.head())
    print("expensesDF")
    print(expensesDF)
    print("categories")
    #print(categoryList.head())
    """


if __name__ == "__main__":
    main()
